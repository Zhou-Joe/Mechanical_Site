import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.db import models
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .models import ECRRecord


def ecr_list(request):
    """ECR记录列表页"""
    records = ECRRecord.objects.all()
    
    # 搜索功能
    search_query = request.GET.get('q')
    if search_query:
        records = records.filter(
            models.Q(title__icontains=search_query) |
            models.Q(requestor__icontains=search_query) |
            models.Q(doc_number__icontains=search_query) |
            models.Q(attraction_or_location__icontains=search_query)
        )
    
    # 分页
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'ecr/ecr_list.html', context)


def ecr_detail(request, pk):
    """ECR记录详情页"""
    record = get_object_or_404(ECRRecord, pk=pk)
    context = {
        'record': record,
    }
    return render(request, 'ecr/ecr_detail.html', context)


def parse_date(date_str):
    """解析日期字符串"""
    if not date_str:
        return None
    
    date_formats = [
        '%Y.%m.%d',
        '%Y-%m-%d',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y/%m/%d',
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    
    return None


def extract_form_number(filename):
    """从文件名中提取表格序号"""
    # 尝试多种模式匹配
    patterns = [
        r'#\s*(\d+)',           # #127 或 # 127
        r'[_\s]#?(\d+)',        # _127 或 _#127
        r'Form[_\s]#?(\d+)',    # Form_127 或 Form #127
        r'(\d+)',               # 任何数字
    ]
    for pattern in patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def import_ecr_from_excel(file_path):
    """从Excel文件导入ECR数据"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 提取表单号（从文件名）
    filename = os.path.basename(file_path)
    form_number = extract_form_number(filename)
    
    # 如果文件名提取失败，尝试从Excel内容读取
    if not form_number:
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for idx, cell_value in enumerate(row):
                if cell_value and 'Form #' in str(cell_value):
                    # 查找同一行的下一个单元格
                    if idx + 1 < len(row) and row[idx + 1]:
                        try:
                            form_number = int(row[idx + 1])
                            break
                        except (ValueError, TypeError):
                            pass
            if form_number:
                break
    
    if not form_number:
        return None, "无法从文件名或文件内容提取表单号"
    
    # 检查是否已存在
    if ECRRecord.objects.filter(form_number=form_number).exists():
        return None, f"表单 #{form_number} 已存在"
    
    # 提取数据
    data = {'form_number': form_number}
    
    # 使用固定位置读取数据（基于实际Excel结构）
    # Row 2: Form #, Progress, Completion Date
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if row2:
        r2 = row2[0]
        if len(r2) > 2 and r2[2]:
            try:
                data['form_number'] = int(r2[2])
            except:
                data['form_number'] = form_number
        if len(r2) > 4 and r2[4]:
            data['progress'] = str(r2[4]).strip()
        if len(r2) > 6 and r2[6]:
            data['completion_date'] = parse_date(r2[6])
    
    # Row 5: Requestor, Contact
    row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))
    if row5:
        r5 = row5[0]
        if len(r5) > 2 and r5[2]:
            data['requestor'] = str(r5[2]).strip()
        if len(r5) > 6 and r5[6]:
            data['contact'] = str(r5[6]).strip()
    
    # Row 6: Request Date, Expected Completion Date
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))
    if row6:
        r6 = row6[0]
        if len(r6) > 2 and r6[2]:
            data['request_date'] = parse_date(r6[2])
        if len(r6) > 5 and r6[5]:
            data['expected_completion_date'] = parse_date(r6[5])
    
    # Row 7: Doc #, REV #
    row7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))
    if row7:
        r7 = row7[0]
        if len(r7) > 2 and r7[2]:
            data['doc_number'] = str(r7[2]).strip()
        if len(r7) > 5 and r7[5]:
            data['rev_number'] = str(r7[5]).strip()
    
    # Row 8: Title
    row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))
    if row8:
        r8 = row8[0]
        if len(r8) > 3 and r8[3]:
            data['title'] = str(r8[3]).strip()
    
    # Row 9: Attraction or Location
    row9 = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    if row9:
        r9 = row9[0]
        if len(r9) > 3 and r9[3]:
            data['attraction_or_location'] = str(r9[3]).strip()
    
    # Row 10: Engineering Change Type
    row10 = list(ws.iter_rows(min_row=10, max_row=10, values_only=True))
    if row10:
        r10 = row10[0]
        for i in range(2, 6):
            if len(r10) > i and r10[i]:
                val = str(r10[i]).strip()
                if val and val != 'None':
                    data['engineering_change_type'] = val
                    break
    
    # Row 12: Detail Description (Row 11是标签，Row 12是内容)
    row12 = list(ws.iter_rows(min_row=12, max_row=12, values_only=True))
    if row12:
        r12 = row12[0]
        if len(r12) > 1 and r12[1]:
            data['detail_description'] = str(r12[1]).strip()
    
    # Row 18-19: Before Change and After Change
    # 先找到标签行
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        for col_idx, cell_value in enumerate(row):
            if not cell_value:
                continue
            cell_str = str(cell_value).strip()
            
            # Before Change (变更前) - 内容在下一行第2列(B列)
            if 'Before Change' in cell_str or '变更前' in cell_str:
                next_row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))
                if next_row and len(next_row[0]) > 1 and next_row[0][1]:
                    data['before_change'] = str(next_row[0][1]).strip()
            
            # After Change (变更后) - 内容在下一行第4列(D列)
            elif 'After Change' in cell_str or '变更后' in cell_str:
                next_row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))
                if next_row and len(next_row[0]) > 3 and next_row[0][3]:
                    data['after_change'] = str(next_row[0][3]).strip()
            
            # Justification (变更理由) - 内容在下一行第2列(B列)
            elif 'Justification' in cell_str or '变更理由' in cell_str:
                next_row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))
                if next_row and len(next_row[0]) > 1 and next_row[0][1]:
                    data['justification'] = str(next_row[0][1]).strip()
    
    # 解析审批信息 - LOB Manager, QE Engineer, ME Engineer
    # 这些通常在20-30行之间
    for row_idx, row in enumerate(ws.iter_rows(min_row=20, max_row=40, values_only=True), 20):
        if not row or len(row) < 2:
            continue
        
        row_str = ' '.join([str(cell).strip() if cell else '' for cell in row[:3]])
        
        # LOB Manager
        if 'LOB Manager' in row_str:
            if len(row) > 1 and row[1]:
                data['lob_manager_name'] = str(row[1]).strip()
            # 查找下一行或同行的审批信息
            for check_row_idx in [row_idx, row_idx + 1]:
                check_row_list = list(ws.iter_rows(min_row=check_row_idx, max_row=check_row_idx, values_only=True))
                if check_row_list:
                    check_row = check_row_list[0]
                    for i, cell in enumerate(check_row):
                        if cell and 'Approval' in str(cell):
                            if i + 1 < len(check_row) and check_row[i + 1]:
                                data['lob_manager_approval'] = str(check_row[i + 1]).strip()
                        if cell and 'Review Date' in str(cell):
                            if i + 1 < len(check_row) and check_row[i + 1]:
                                data['lob_manager_review_date'] = parse_date(check_row[i + 1])
        
        # QE Engineer
        elif 'QE Engineer' in row_str:
            if len(row) > 1 and row[1]:
                data['qe_engineer_name'] = str(row[1]).strip()
            for check_row_idx in [row_idx, row_idx + 1]:
                check_row_list = list(ws.iter_rows(min_row=check_row_idx, max_row=check_row_idx, values_only=True))
                if check_row_list:
                    check_row = check_row_list[0]
                    for i, cell in enumerate(check_row):
                        if cell and 'Approval' in str(cell):
                            if i + 1 < len(check_row) and check_row[i + 1]:
                                data['qe_engineer_approval'] = str(check_row[i + 1]).strip()
                        if cell and 'Review Date' in str(cell):
                            if i + 1 < len(check_row) and check_row[i + 1]:
                                data['qe_engineer_review_date'] = parse_date(check_row[i + 1])
        
        # ME Engineer
        elif 'ME Engineer' in row_str:
            if len(row) > 1 and row[1]:
                data['me_engineer_name'] = str(row[1]).strip()
            for check_row_idx in [row_idx, row_idx + 1]:
                check_row_list = list(ws.iter_rows(min_row=check_row_idx, max_row=check_row_idx, values_only=True))
                if check_row_list:
                    check_row = check_row_list[0]
                    for i, cell in enumerate(check_row):
                        if cell and 'Approval' in str(cell):
                            if i + 1 < len(check_row) and check_row[i + 1]:
                                data['me_engineer_approval'] = str(check_row[i + 1]).strip()
                        if cell and 'Review Date' in str(cell):
                            if i + 1 < len(check_row) and check_row[i + 1]:
                                data['me_engineer_review_date'] = parse_date(check_row[i + 1])
    
    # 创建记录
    try:
        record = ECRRecord.objects.create(**data)
        return record, None
    except Exception as e:
        return None, str(e)


def ecr_import(request):
    """处理ECR Excel文件导入"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # 检查文件类型
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, '请上传Excel文件(.xlsx或.xls)')
            return redirect('ecr:ecr_list')
        
        # 保存上传的文件
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        record = None
        error = None
        
        try:
            # 导入数据
            record, error = import_ecr_from_excel(tmp_file_path)
        except Exception as e:
            error = str(e)
        finally:
            # 删除临时文件
            try:
                os.unlink(tmp_file_path)
            except:
                pass
        
        # 显示结果消息
        if record:
            messages.success(request, f'成功导入 ECR #{record.form_number}: {record.title or "Untitled"}')
        else:
            messages.error(request, f'导入失败: {error}')
        
        return redirect('ecr:ecr_list')
    
    return redirect('ecr:ecr_list')


from datetime import datetime, timedelta, date

def ecr_create(request):
    """创建新的ECR记录"""
    # 获取最大的form_number并+1作为默认值
    from django.db.models import Max
    max_form_number = ECRRecord.objects.aggregate(Max('form_number'))['form_number__max']
    next_form_number = (max_form_number or 0) + 1
    
    if request.method == 'POST':
        # 获取表单数据
        form_number = request.POST.get('form_number')
        
        # 检查表单号是否已存在
        if ECRRecord.objects.filter(form_number=form_number).exists():
            messages.error(request, f'表单号 #{form_number} 已存在')
            return render(request, 'ecr/ecr_form.html', {
                'request': request,
                'next_form_number': next_form_number
            })
        
        # 计算默认的预期完成日期（1个月后）
        default_expected_date = (datetime.now() + timedelta(days=30)).date()
        expected_completion_date = request.POST.get('expected_completion_date')
        if not expected_completion_date:
            expected_completion_date = default_expected_date
        
        # 创建记录
        try:
            record = ECRRecord.objects.create(
                form_number=form_number,
                progress=request.POST.get('progress', 'In Progress'),
                completion_date=request.POST.get('completion_date') or None,
                requestor=request.POST.get('requestor', ''),
                contact=request.POST.get('contact', ''),
                request_date=request.POST.get('request_date') or None,
                expected_completion_date=expected_completion_date,
                doc_number=request.POST.get('doc_number', ''),
                rev_number=request.POST.get('rev_number', ''),
                title=request.POST.get('title', ''),
                attraction_or_location=request.POST.get('attraction_or_location', ''),
                engineering_change_type=request.POST.get('engineering_change_type', ''),
                detail_description=request.POST.get('detail_description', ''),
                before_change=request.POST.get('before_change', ''),
                after_change=request.POST.get('after_change', ''),
                justification=request.POST.get('justification', ''),
                # LOB Manager
                lob_manager_name=request.POST.get('lob_manager_name', ''),
                lob_manager_approval=request.POST.get('lob_manager_approval', ''),
                lob_manager_review_date=request.POST.get('lob_manager_review_date') or None,
                lob_manager_comments=request.POST.get('lob_manager_comments', ''),
                # QE Engineer - 仅管理员可设置
                qe_engineer_name=request.POST.get('qe_engineer_name', '') if request.user.is_staff else '',
                qe_engineer_approval=request.POST.get('qe_engineer_approval', '') if request.user.is_staff else '',
                qe_engineer_review_date=(request.POST.get('qe_engineer_review_date') or None) if request.user.is_staff else None,
                qe_engineer_comments=request.POST.get('qe_engineer_comments', '') if request.user.is_staff else '',
                # ME Engineer - 仅管理员可设置
                me_engineer_name=request.POST.get('me_engineer_name', '') if request.user.is_staff else '',
                me_engineer_approval=request.POST.get('me_engineer_approval', '') if request.user.is_staff else '',
                me_engineer_review_date=(request.POST.get('me_engineer_review_date') or None) if request.user.is_staff else None,
                me_engineer_comments=request.POST.get('me_engineer_comments', '') if request.user.is_staff else '',
                # 图片
                before_change_image=request.FILES.get('before_change_image'),
                after_change_image=request.FILES.get('after_change_image'),
            )
            messages.success(request, f'成功创建 ECR #{record.form_number}')
            return redirect('ecr:ecr_list')
        except Exception as e:
            messages.error(request, f'创建失败: {str(e)}')
    
    # 传递今天的日期和默认预期完成日期给模板
    today = date.today().strftime('%Y-%m-%d')
    default_expected_date = (date.today() + timedelta(days=30)).strftime('%Y-%m-%d')
    return render(request, 'ecr/ecr_form.html', {
        'today': today,
        'default_expected_date': default_expected_date,
        'next_form_number': next_form_number
    })


def ecr_edit(request, pk):
    """编辑ECR记录"""
    record = get_object_or_404(ECRRecord, pk=pk)
    
    if request.method == 'POST':
        # 获取表单数据
        form_number = request.POST.get('form_number')
        
        # 检查表单号是否被其他记录使用
        if form_number and int(form_number) != record.form_number:
            if ECRRecord.objects.filter(form_number=form_number).exists():
                messages.error(request, f'表单号 #{form_number} 已被其他记录使用')
                return render(request, 'ecr/ecr_form.html', {'record': record})
        
        # 更新记录
        try:
            record.form_number = form_number or record.form_number
            record.progress = request.POST.get('progress', '')
            record.completion_date = request.POST.get('completion_date') or None
            record.requestor = request.POST.get('requestor', '')
            record.contact = request.POST.get('contact', '')
            record.request_date = request.POST.get('request_date') or None
            record.expected_completion_date = request.POST.get('expected_completion_date') or None
            record.doc_number = request.POST.get('doc_number', '')
            record.rev_number = request.POST.get('rev_number', '')
            record.title = request.POST.get('title', '')
            record.attraction_or_location = request.POST.get('attraction_or_location', '')
            record.engineering_change_type = request.POST.get('engineering_change_type', '')
            record.detail_description = request.POST.get('detail_description', '')
            record.before_change = request.POST.get('before_change', '')
            record.after_change = request.POST.get('after_change', '')
            record.justification = request.POST.get('justification', '')
            # LOB Manager
            record.lob_manager_name = request.POST.get('lob_manager_name', '')
            record.lob_manager_approval = request.POST.get('lob_manager_approval', '')
            record.lob_manager_review_date = request.POST.get('lob_manager_review_date') or None
            record.lob_manager_comments = request.POST.get('lob_manager_comments', '')
            # QE Engineer - 仅管理员可编辑
            if request.user.is_staff:
                record.qe_engineer_name = request.POST.get('qe_engineer_name', '')
                record.qe_engineer_approval = request.POST.get('qe_engineer_approval', '')
                record.qe_engineer_review_date = request.POST.get('qe_engineer_review_date') or None
                record.qe_engineer_comments = request.POST.get('qe_engineer_comments', '')
            # ME Engineer - 仅管理员可编辑
            if request.user.is_staff:
                record.me_engineer_name = request.POST.get('me_engineer_name', '')
                record.me_engineer_approval = request.POST.get('me_engineer_approval', '')
                record.me_engineer_review_date = request.POST.get('me_engineer_review_date') or None
                record.me_engineer_comments = request.POST.get('me_engineer_comments', '')
            
            # 图片
            if request.FILES.get('before_change_image'):
                record.before_change_image = request.FILES.get('before_change_image')
            if request.FILES.get('after_change_image'):
                record.after_change_image = request.FILES.get('after_change_image')
            
            record.save()
            messages.success(request, f'成功更新 ECR #{record.form_number}')
            return redirect('ecr:ecr_detail', pk=record.pk)
        except Exception as e:
            messages.error(request, f'更新失败: {str(e)}')
    
    # 传递今天的日期和默认预期完成日期给模板
    from datetime import date, timedelta
    from django.db.models import Max
    today = date.today().strftime('%Y-%m-%d')
    default_expected_date = (date.today() + timedelta(days=30)).strftime('%Y-%m-%d')
    max_form_number = ECRRecord.objects.aggregate(Max('form_number'))['form_number__max']
    next_form_number = (max_form_number or 0) + 1
    return render(request, 'ecr/ecr_form.html', {
        'record': record,
        'today': today,
        'default_expected_date': default_expected_date,
        'next_form_number': next_form_number
    })


def export_ecr_to_excel(record, output_path):
    """将ECR记录导出为Excel文件"""
    from openpyxl.cell.cell import MergedCell
    
    # 加载模板
    template_path = Path(__file__).resolve().parent.parent / 'ECR_Record' / 'Engineering Change Request Form_#127.xlsx'
    
    if not template_path.exists():
        # 如果模板不存在，尝试查找任何ECR文件作为模板
        ecr_dir = Path(__file__).resolve().parent.parent / 'ECR_Record'
        template_files = list(ecr_dir.glob('*.xlsx'))
        if template_files:
            template_path = template_files[0]
        else:
            return None, "找不到Excel模板文件"
    
    # 加载工作簿
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 填充数据到对应单元格
    # 使用字典记录每个标签的位置，避免重复匹配
    filled_cells = set()
    
    # 先找到Progress和Completion Date标签的位置，然后填入对应值
    progress_row = None
    progress_col = None
    completion_date_row = None
    completion_date_col = None
    
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            # 跳过合并单元格
            if isinstance(cell, MergedCell):
                continue
            
            if not cell.value:
                continue
            
            cell_str = str(cell.value).strip()
            
            # 记录Progress标签位置
            if 'Progress' in cell_str and '状态' in cell_str:
                progress_row = cell.row
                progress_col = cell.column
            
            # 记录Completion Date标签位置（排除Expected Completion Date）
            elif ('Completion Date' in cell_str or '完成日期' in cell_str) and 'Expected' not in cell_str and '预期' not in cell_str:
                completion_date_row = cell.row
                completion_date_col = cell.column
    
    # 直接填入数据到指定位置
    # Progress -> 行2 列E (第5列)
    if record.progress:
        target_cell = ws.cell(row=2, column=5)  # E2
        if not isinstance(target_cell, MergedCell):
            target_cell.value = record.progress
    
    # Completion Date -> 行2 列G (第7列)
    if record.completion_date:
        target_cell = ws.cell(row=2, column=7)  # G2
        if not isinstance(target_cell, MergedCell):
            target_cell.value = record.completion_date.strftime('%Y.%m.%d')
    
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            # 跳过合并单元格
            if isinstance(cell, MergedCell):
                continue
            
            if not cell.value:
                continue
            
            cell_str = str(cell.value).strip()
            cell_key = (cell.row, cell.column)
            
            if cell_key in filled_cells:
                continue
            
            # Form # (表格序号) - 匹配 "Form # / 表格序号"
            if cell_str == 'Form # / 表格序号' or ('Form #' in cell_str and '表格序号' in cell_str):
                target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if not isinstance(target_cell, MergedCell):
                    target_cell.value = record.form_number
                    filled_cells.add(cell_key)
            
            # Requestor (需求者)
            elif 'Requestor' in cell_str or '需求者' in cell_str:
                if record.requestor:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.requestor
            
            # Contact (联系方式)
            elif 'Contact' in cell_str or '联系方式' in cell_str:
                if record.contact:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.contact
            
            # Request Date (请求日期)
            elif 'Request Date' in cell_str or '请求日期' in cell_str:
                if record.request_date:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.request_date.strftime('%Y.%m.%d')
            
            # Expected Completion Date (预期完成日期)
            elif 'Expected Completion Date' in cell_str or '预期完成日期' in cell_str:
                if record.expected_completion_date:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.expected_completion_date.strftime('%Y.%m.%d')
            
            # Doc # (文件号)
            elif 'Doc #' in cell_str or '文件号' in cell_str:
                if record.doc_number:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.doc_number
            
            # REV # (版本号)
            elif 'REV #' in cell_str or '版本号' in cell_str:
                if record.rev_number:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.rev_number
            
            # Title (标题)
            elif 'Title' in cell_str or '标题' in cell_str:
                if record.title:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.title
            
            # Attraction or Location (景点或地点)
            elif 'Attraction' in cell_str or '景点' in cell_str:
                if record.attraction_or_location:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.attraction_or_location
            
            # Engineering Change Type (文件类型)
            elif 'Engineering Change Type' in cell_str or '文件类型' in cell_str:
                if record.engineering_change_type:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.engineering_change_type
            
            # Detail Description (详细描述)
            elif 'Detail Description' in cell_str or '详细描述' in cell_str:
                if record.detail_description:
                    target_cell = ws.cell(row=cell.row + 1, column=cell.column)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.detail_description
            
            # Before Change (变更前)
            elif 'Before Change' in cell_str or '变更前' in cell_str:
                if record.before_change:
                    target_cell = ws.cell(row=cell.row + 1, column=cell.column)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.before_change
            
            # After Change (变更后)
            elif 'After Change' in cell_str or '变更后' in cell_str:
                if record.after_change:
                    target_cell = ws.cell(row=cell.row + 1, column=cell.column)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.after_change
            
            # Justification (变更理由)
            elif 'Justification' in cell_str or '变更理由' in cell_str:
                if record.justification:
                    target_cell = ws.cell(row=cell.row + 1, column=cell.column)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.justification
            
            # LOB Manager
            elif 'LOB Manager' in cell_str:
                if record.lob_manager_name:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.lob_manager_name
            
            # QE Engineer
            elif 'QE Engineer' in cell_str:
                if record.qe_engineer_name:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.qe_engineer_name
            
            # ME Engineer
            elif 'ME Engineer' in cell_str:
                if record.me_engineer_name:
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = record.me_engineer_name
            
            # Approval (Y/N)
            elif 'Approval (Y/N)' in cell_str:
                # 需要根据上下文判断是哪个角色的审批
                # 这里简化处理，在找到角色名后填充
                pass
            
            # Review Date
            elif 'Review Date' in cell_str:
                # 需要根据上下文判断是哪个角色的审批日期
                pass
    
    # 第二遍遍历填充审批信息和日期
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if not cell.value:
                continue
            
            cell_str = str(cell.value).strip()
            
            # 查找LOB Manager的审批行
            if 'LOB Manager' in cell_str:
                # 在当前行和下一行查找Approval和Review Date
                for r in range(cell.row, min(cell.row + 3, 50)):
                    for c in range(1, 10):
                        check_cell = ws.cell(row=r, column=c)
                        if isinstance(check_cell, MergedCell):
                            continue
                        if check_cell.value and 'Approval (Y/N)' in str(check_cell.value):
                            target = ws.cell(row=r, column=c + 1)
                            if not isinstance(target, MergedCell) and record.lob_manager_approval:
                                target.value = record.lob_manager_approval
                        if check_cell.value and 'Review Date' in str(check_cell.value):
                            target = ws.cell(row=r, column=c + 1)
                            if not isinstance(target, MergedCell) and record.lob_manager_review_date:
                                target.value = record.lob_manager_review_date.strftime('%Y.%m.%d')
            
            # 查找QE Engineer的审批行
            elif 'QE Engineer' in cell_str:
                for r in range(cell.row, min(cell.row + 3, 50)):
                    for c in range(1, 10):
                        check_cell = ws.cell(row=r, column=c)
                        if isinstance(check_cell, MergedCell):
                            continue
                        if check_cell.value and 'Approval (Y/N)' in str(check_cell.value):
                            target = ws.cell(row=r, column=c + 1)
                            if not isinstance(target, MergedCell) and record.qe_engineer_approval:
                                target.value = record.qe_engineer_approval
                        if check_cell.value and 'Review Date' in str(check_cell.value):
                            target = ws.cell(row=r, column=c + 1)
                            if not isinstance(target, MergedCell) and record.qe_engineer_review_date:
                                target.value = record.qe_engineer_review_date.strftime('%Y.%m.%d')
            
            # 查找ME Engineer的审批行
            elif 'ME Engineer' in cell_str:
                for r in range(cell.row, min(cell.row + 3, 50)):
                    for c in range(1, 10):
                        check_cell = ws.cell(row=r, column=c)
                        if isinstance(check_cell, MergedCell):
                            continue
                        if check_cell.value and 'Approval (Y/N)' in str(check_cell.value):
                            target = ws.cell(row=r, column=c + 1)
                            if not isinstance(target, MergedCell) and record.me_engineer_approval:
                                target.value = record.me_engineer_approval
                        if check_cell.value and 'Review Date' in str(check_cell.value):
                            target = ws.cell(row=r, column=c + 1)
                            if not isinstance(target, MergedCell) and record.me_engineer_review_date:
                                target.value = record.me_engineer_review_date.strftime('%Y.%m.%d')
    
    # 插入图片
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    import io
    
    # 查找Before Change和After Change图片位置
    before_change_img_row = None
    before_change_img_col = None
    after_change_img_row = None
    after_change_img_col = None
    
    for row in ws.iter_rows(min_row=1, max_row=100):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if not cell.value:
                continue
            cell_str = str(cell.value).strip()
            
            # 查找Before Change图片位置（通常在"Before Change"标签下方）
            if 'Before Change' in cell_str or '变更前' in cell_str:
                before_change_img_row = cell.row + 2  # 标签下方2行
                before_change_img_col = cell.column
            
            # 查找After Change图片位置
            elif 'After Change' in cell_str or '变更后' in cell_str:
                after_change_img_row = cell.row + 2  # 标签下方2行
                after_change_img_col = cell.column
    
    # 插入Before Change图片
    if record.before_change_image and record.before_change_image.path:
        try:
            # 打开图片并调整大小
            img_path = record.before_change_image.path
            pil_img = PILImage.open(img_path)
            
            # 调整图片大小（最大宽度400像素）
            max_width = 400
            ratio = max_width / pil_img.width
            new_height = int(pil_img.height * ratio)
            pil_img = pil_img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)
            
            # 保存到内存
            img_buffer = io.BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # 创建Excel图片对象
            xl_img = XLImage(img_buffer)
            
            # 设置图片位置
            if before_change_img_row and before_change_img_col:
                cell_addr = f'{chr(64 + before_change_img_col)}{before_change_img_row}'
                ws.add_image(xl_img, cell_addr)
            else:
                # 默认位置
                ws.add_image(xl_img, 'B25')
        except Exception as e:
            print(f"插入Before Change图片失败: {e}")
    
    # 插入After Change图片
    if record.after_change_image and record.after_change_image.path:
        try:
            # 打开图片并调整大小
            img_path = record.after_change_image.path
            pil_img = PILImage.open(img_path)
            
            # 调整图片大小（最大宽度400像素）
            max_width = 400
            ratio = max_width / pil_img.width
            new_height = int(pil_img.height * ratio)
            pil_img = pil_img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)
            
            # 保存到内存
            img_buffer = io.BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # 创建Excel图片对象
            xl_img = XLImage(img_buffer)
            
            # 设置图片位置
            if after_change_img_row and after_change_img_col:
                cell_addr = f'{chr(64 + after_change_img_col)}{after_change_img_row}'
                ws.add_image(xl_img, cell_addr)
            else:
                # 默认位置
                ws.add_image(xl_img, 'H25')
        except Exception as e:
            print(f"插入After Change图片失败: {e}")
    
    # 保存文件
    wb.save(output_path)
    return output_path, None


def ecr_export(request, pk):
    """导出单个ECR记录为Excel文件"""
    record = get_object_or_404(ECRRecord, pk=pk)
    
    # 创建临时文件
    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp_file:
        tmp_file_path = tmp_file.name
    
    try:
        # 导出数据
        output_path, error = export_ecr_to_excel(record, tmp_file_path)
        
        if error:
            messages.error(request, f'导出失败: {error}')
            return redirect('ecr:ecr_list')
        
        # 读取文件并返回下载响应
        with open(output_path, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="ECR_#{record.form_number}.xlsx"'
            return response
            
    except Exception as e:
        messages.error(request, f'导出失败: {str(e)}')
        return redirect('ecr:ecr_list')
    finally:
        # 删除临时文件
        try:
            os.unlink(tmp_file_path)
        except:
            pass
