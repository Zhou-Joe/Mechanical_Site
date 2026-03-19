"""
ECR Excel 数据导入脚本
"""
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl

# 设置 Django 环境
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'oil_inspection.settings')
import django
django.setup()

from ecr.models import ECRRecord


def parse_date(date_str):
    """解析日期字符串"""
    if not date_str:
        return None
    
    # 尝试多种日期格式
    date_formats = [
        '%Y.%m.%d',
        '%Y-%m-%d',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y/%m/%d',
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    
    return None


def extract_form_number(filename):
    """从文件名中提取表格序号"""
    # 匹配 #数字 或 _数字 格式
    match = re.search(r'[#_](\d+)', filename)
    if match:
        return int(match.group(1))
    return None


def import_ecr_from_excel(file_path):
    """
    从Excel文件导入ECR数据
    
    参数:
        file_path: Excel文件路径
    """
    print(f"正在导入: {file_path}")
    
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 提取表单号（从文件名）
    filename = os.path.basename(file_path)
    form_number = extract_form_number(filename)
    
    if not form_number:
        print(f"警告: 无法从文件名提取表单号: {filename}")
        return None
    
    # 检查是否已存在
    if ECRRecord.objects.filter(form_number=form_number).exists():
        print(f"表单 #{form_number} 已存在，跳过")
        return None
    
    # 提取数据
    data = {
        'form_number': form_number,
    }
    
    # 先直接从固定位置读取Progress和Completion Date
    # Progress在E2 (第2行第5列)
    progress_cell = ws.cell(row=2, column=5)
    if progress_cell.value:
        data['progress'] = str(progress_cell.value).strip()
    
    # Completion Date在G2 (第2行第7列)
    completion_date_cell = ws.cell(row=2, column=7)
    if completion_date_cell.value:
        data['completion_date'] = parse_date(completion_date_cell.value)
    
    # 遍历所有单元格查找其他数据
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                
                # 跳过已经处理的Progress和Completion Date标签
                if ('Progress' in cell_value and '状态' in cell_value) or \
                   (('Completion Date' in cell_value or '完成日期' in cell_value) and 
                    'Expected' not in cell_value and '预期' not in cell_value):
                    continue
                
                elif 'Requestor' in cell_value or '需求者' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['requestor'] = str(next_cell.value).strip()
                
                elif 'Contact' in cell_value or '联系方式' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['contact'] = str(next_cell.value).strip()
                
                elif 'Request Date' in cell_value or '请求日期' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['request_date'] = parse_date(next_cell.value)
                
                elif 'Expected Completion Date' in cell_value or '预期完成日期' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['expected_completion_date'] = parse_date(next_cell.value)
                
                elif 'Doc #' in cell_value or '文件号' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['doc_number'] = str(next_cell.value).strip()
                
                elif 'REV #' in cell_value or '版本号' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['rev_number'] = str(next_cell.value).strip()
                
                elif 'Title' in cell_value or '标题' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['title'] = str(next_cell.value).strip()
                
                elif 'Attraction or Location' in cell_value or '景点或地点' in cell_value:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    if next_cell.value:
                        data['attraction_or_location'] = str(next_cell.value).strip()
                
                elif 'Engineering Change Type' in cell_value or '文件类型' in cell_value:
                    # 查找类型值（可能在后面的列）
                    for col_offset in range(2, 6):
                        type_cell = ws.cell(row=cell.row, column=cell.column + col_offset)
                        if type_cell.value and ('MM' in str(type_cell.value) or 
                                               '维护手册' in str(type_cell.value) or
                                               'OM' in str(type_cell.value) or
                                               '操作手册' in str(type_cell.value)):
                            data['engineering_change_type'] = str(type_cell.value).strip()
                            break
                
                elif 'Detail Description' in cell_value or '详细描述' in cell_value:
                    # 详细描述在下一行
                    next_row = ws.cell(row=cell.row + 1, column=cell.column)
                    if next_row.value:
                        data['detail_description'] = str(next_row.value).strip()
                
                elif 'Before Change' in cell_value or '变更前' in cell_value:
                    # 变更前内容在下一行
                    next_row = ws.cell(row=cell.row + 1, column=cell.column)
                    if next_row.value:
                        data['before_change'] = str(next_row.value).strip()
                
                elif 'After Change' in cell_value or '变更后' in cell_value:
                    # 变更后内容在下一行，同一列
                    next_row = ws.cell(row=cell.row + 1, column=cell.column)
                    if next_row.value:
                        data['after_change'] = str(next_row.value).strip()
                
                elif 'Justification' in cell_value or '变更理由' in cell_value:
                    # 变更理由在下一行
                    next_row = ws.cell(row=cell.row + 1, column=cell.column)
                    if next_row.value:
                        data['justification'] = str(next_row.value).strip()
                
                # 审批信息处理
                elif 'LOB Manager' in cell_value:
                    # 在LOB Manager行附近查找Name、Approval、Review Date
                    for r in range(cell.row, min(cell.row + 3, 50)):
                        for c in range(1, 10):
                            check_cell = ws.cell(row=r, column=c)
                            if not check_cell.value:
                                continue
                            check_str = str(check_cell.value).strip()
                            
                            # Name
                            if check_str == 'Name' or check_str == '名称':
                                name_cell = ws.cell(row=r, column=c + 1)
                                if name_cell.value:
                                    data['lob_manager_name'] = str(name_cell.value).strip()
                            
                            # Approval (Y/N)
                            if 'Approval (Y/N)' in check_str or '审批 (Y/N)' in check_str:
                                approval_cell = ws.cell(row=r, column=c + 1)
                                if approval_cell.value:
                                    data['lob_manager_approval'] = str(approval_cell.value).strip()
                            
                            # Review Date
                            if 'Review Date' in check_str or '审核日期' in check_str:
                                date_cell = ws.cell(row=r, column=c + 1)
                                if date_cell.value:
                                    data['lob_manager_review_date'] = parse_date(date_cell.value)
                            
                            # Comments
                            if 'Comments' in check_str or '备注' in check_str:
                                comments_cell = ws.cell(row=r, column=c + 1)
                                if comments_cell.value:
                                    data['lob_manager_comments'] = str(comments_cell.value).strip()
                
                elif 'QE Engineer' in cell_value:
                    # QE Engineer审批信息
                    for r in range(cell.row, min(cell.row + 3, 50)):
                        for c in range(1, 10):
                            check_cell = ws.cell(row=r, column=c)
                            if not check_cell.value:
                                continue
                            check_str = str(check_cell.value).strip()
                            
                            if check_str == 'Name' or check_str == '名称':
                                name_cell = ws.cell(row=r, column=c + 1)
                                if name_cell.value:
                                    data['qe_engineer_name'] = str(name_cell.value).strip()
                            
                            if 'Approval (Y/N)' in check_str or '审批 (Y/N)' in check_str:
                                approval_cell = ws.cell(row=r, column=c + 1)
                                if approval_cell.value:
                                    data['qe_engineer_approval'] = str(approval_cell.value).strip()
                            
                            if 'Review Date' in check_str or '审核日期' in check_str:
                                date_cell = ws.cell(row=r, column=c + 1)
                                if date_cell.value:
                                    data['qe_engineer_review_date'] = parse_date(date_cell.value)
                            
                            if 'Comments' in check_str or '备注' in check_str:
                                comments_cell = ws.cell(row=r, column=c + 1)
                                if comments_cell.value:
                                    data['qe_engineer_comments'] = str(comments_cell.value).strip()
                
                elif 'ME Engineer' in cell_value:
                    # ME Engineer审批信息
                    for r in range(cell.row, min(cell.row + 3, 50)):
                        for c in range(1, 10):
                            check_cell = ws.cell(row=r, column=c)
                            if not check_cell.value:
                                continue
                            check_str = str(check_cell.value).strip()
                            
                            if check_str == 'Name' or check_str == '名称':
                                name_cell = ws.cell(row=r, column=c + 1)
                                if name_cell.value:
                                    data['me_engineer_name'] = str(name_cell.value).strip()
                            
                            if 'Approval (Y/N)' in check_str or '审批 (Y/N)' in check_str:
                                approval_cell = ws.cell(row=r, column=c + 1)
                                if approval_cell.value:
                                    data['me_engineer_approval'] = str(approval_cell.value).strip()
                            
                            if 'Review Date' in check_str or '审核日期' in check_str:
                                date_cell = ws.cell(row=r, column=c + 1)
                                if date_cell.value:
                                    data['me_engineer_review_date'] = parse_date(date_cell.value)
                            
                            if 'Comments' in check_str or '备注' in check_str:
                                comments_cell = ws.cell(row=r, column=c + 1)
                                if comments_cell.value:
                                    data['me_engineer_comments'] = str(comments_cell.value).strip()
    
    # 创建记录
    try:
        record = ECRRecord.objects.create(**data)
        print(f"成功导入 ECR #{form_number}: {data.get('title', 'Untitled')}")
        return record
    except Exception as e:
        print(f"导入失败 ECR #{form_number}: {e}")
        return None


def import_all_ecr_files(directory):
    """
    导入目录中的所有ECR Excel文件
    
    参数:
        directory: ECR_Record 目录路径
    """
    directory = Path(directory)
    
    if not directory.exists():
        print(f"目录不存在: {directory}")
        return
    
    # 查找所有Excel文件
    excel_files = list(directory.glob('*.xlsx')) + list(directory.glob('*.xls'))
    
    if not excel_files:
        print(f"目录中没有Excel文件: {directory}")
        return
    
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    success_count = 0
    for file_path in excel_files:
        if import_ecr_from_excel(file_path):
            success_count += 1
    
    print(f"\n导入完成: {success_count}/{len(excel_files)} 个文件成功导入")


if __name__ == '__main__':
    # 导入单个示例文件
    sample_file = Path(__file__).resolve().parent.parent / 'ECR_Record' / 'Engineering Change Request Form_#127.xlsx'
    
    if sample_file.exists():
        import_ecr_from_excel(sample_file)
    else:
        # 导入整个目录
        ecr_directory = Path(__file__).resolve().parent.parent / 'ECR_Record'
        import_all_ecr_files(ecr_directory)
